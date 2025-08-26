import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import re
import numpy as np
from openpyxl import load_workbook
import regex 


#获取gaitevent数据
def gaitEventLoader(_Dir):
    wb = load_workbook(_Dir)
    shs = wb.worksheets
    sh = shs[0]
    _eventDic = {}
    _frameDic = {}
    walknum = 1
    while sh.cell(1,walknum).value!=None:
        eventnum = 2 #此处可能为2，需要看导入的xlxs什么样
        _eventList = []
        _frameList = []
        while sh.cell(eventnum,walknum).value!=None:
            _eventList.append(sh.cell(eventnum,walknum).value)
            _frameList.append(sh.cell(eventnum,walknum+1).value)
            eventnum += 1            
        pattern_num = regex.compile(r"(?<=[wW][aA][lL][kK]0*)[1-9]+[0-9]*")
        Key = pattern_num.findall(sh.cell(1,walknum).value)[0]
        _eventDic[Key] = _eventList
        _frameDic[Key] = _frameList
        walknum += 3
    return _eventDic, _frameDic

#获取脑电trigger的事件数据
def OpenEvent():
    # 创建一个隐藏的根窗口
    root = tk.Tk()
    root.withdraw()  # 隐藏根窗口
    
    # 打开文件选择对话框
    file_path = filedialog.askopenfilename(filetypes=[("Event BDF files", "*.bdf")])
    
    if file_path:
        # 读取文件
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            bdf_content = file.read()
        
        pattern1 = r'\+(\d+\.\d+)'  # 匹配+后面跟着的浮点数
        pattern2 = r'\x14(?!\x14)([^\x15\x00]+?)\x14' # 匹配事件名称U+0014Eyes OpenU+0014

        matches1 = re.findall(pattern1, bdf_content)
        matches2 = re.findall(pattern2, bdf_content)
        # 创建结果列表
        time_list = []
        event_list = []
        for match in matches1:
            time = match  # 事件时刻
            time_list.append(time)  # 添加到时间列表
        for match in matches2:
            event = match  # 事件名称
            event_list.append(event)  # 添加到事件列表
        root.destroy()  # 选择文件后销毁根窗口
        return event_list, time_list
    else:
        messagebox.showinfo("提示", "请选择正确的文件")
        root.destroy()  # 如果没有选择文件，销毁根窗口

#打开GE的xlsx文件
def OpenGEfile():
    # 创建一个隐藏的根窗口
    root = tk.Tk()
    root.withdraw()  # 隐藏根窗口
    
    # 打开文件选择对话框
    file_path = filedialog.askopenfilename(filetypes=[("GaitEvent files", "*.xlsx")])
    
    if file_path:
        # 读取文件
        GE_Event_dic, GE_Frame_dic = gaitEventLoader(file_path)
        root.destroy()  # 选择文件后销毁根窗口
        return GE_Event_dic, GE_Frame_dic
    else:
        messagebox.showinfo("提示", "请选择正确的文件")
        root.destroy()  # 如果没有选择文件，销毁根窗口
