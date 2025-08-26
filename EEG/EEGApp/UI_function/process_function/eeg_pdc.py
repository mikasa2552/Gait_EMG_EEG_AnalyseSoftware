import numpy as np
import matplotlib.pyplot as plt
from nilearn import plotting
from .pypdc.analyze_pdc import analyze_pdc
from .pypdc.graph_cal import calculate_graph_metrics
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
import mne
import tkinter as tk
from tkinter import messagebox
from tkinter.filedialog import asksaveasfilename
import warnings
warnings.filterwarnings("ignore")


def get_info(raw):
    """
    获取RAW的基本信息

    -------------

    输出：
        eeg_list,eeg信号，list中每个单元储存一个eeg信号，m*n，m为通道数，n为信号长度

        coordinates_matrix：通道的MNI位置，m*3，m为通道数，3为三维坐标

        channel_labels：通道的名称，为list

        fs：采样频率，默认1000Hz

    """
    eeg_list=[]
    if type(raw) == list:#list形式
        for eeg in raw:
            eeg_list.append(eeg.get_data())
        channel_labels=raw[0].ch_names
        fs=raw[0].info['sfreq']
        
    elif type(raw) == mne.epochs.Epochs:#epochs形式
        n_epochs=len(raw)
        data=raw.get_data()
        for i in range(n_epochs):
            eeg_list.append(data[i])
        channel_labels=raw.ch_names
        fs=raw.info['sfreq']
        
    else:#mne.raw形式
        eeg_list.append(raw.get_data())
        channel_labels=raw.ch_names
        fs=raw.info['sfreq']
    
    file_path = r'EEGApp\UI_function\process_function\10-10channels_loc.xlsx'  # 替换为10-10channels_loc.xlsx的文件路径
    df = pd.read_excel(file_path)

    # 处理坐标数据：只保留前面的数值
    for col in ['X(mm)','Y(mm)','Z(mm)']:
        df[col] = df[col].str.split('±').str[0].astype(float)


    # 构建 n * 3 的矩阵
    coordinates_matrix = np.zeros((len(channel_labels), 3))  # 初始化矩阵

    # 将 Excel 中的 Labels 列转换为大写，以便不区分大小写匹配
    df['Labels'] = df['Labels'].str.upper()

    # 遍历通道名称列表
    for i, ch_name in enumerate(channel_labels):
        # 将通道名称转换为大写
        ch_name_upper = ch_name.upper()
        
        # 找到通道对应的坐标
        if ch_name_upper in df['Labels'].values:
            row = df[df['Labels'] == ch_name_upper]
            coordinates_matrix[i] = row[['X(mm)','Y(mm)','Z(mm)']].values[0]
        else:
            print(f"警告：通道 {ch_name} 未找到对应的坐标！")    
    
    return eeg_list,coordinates_matrix,channel_labels,fs


def PDC_cal(eeg_list,coords,channel_labels,fs=1000,alpha=0.05):
    
    """
    计算EEG信号的PDC

    -------------

    输入：
        eeg_list，eeg信号，list中每个单元储存一个eeg信号，m*n，m为通道数，n为信号长度

        coords：通道的MNI位置，m*3，m为通道数，3为三维坐标

        channel_labels：通道的名称，为list

        fs：采样频率，默认1000Hz

        alpha：进行显著性检验时的水平，默认0.05，即95%的显著性水平

    -------------

    输出：
        wb，包含五个频段各个指标的workbook

    """
    result=analyze_pdc(eeg_list,sampling_rate=fs,alpha=alpha)

    wb = Workbook()

    # 创建一个对齐样式，设置垂直居中对齐
    center_alignment = Alignment(vertical='center')

    # 遍历每个频段
    for band_name, binary_matrix in result.items():

        
        # 计算图论指标
        results = calculate_graph_metrics(binary_matrix['weighted_pdc_matrix'], is_binary=False)
        
        # 创建一个新的 sheet
        ws = wb.create_sheet(title=band_name)
        
        # 写入表头  
        headers = [
            'Channel', 'Node Strength', 'Node Degree', 'Clustering Coefficient',
            'Local Efficiency List', 'Characteristic Path Length', 'Global Efficiency', 'Local Efficiency'
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.alignment = center_alignment  # 设置垂直居中对齐
        
        # 写入多值指标
        node_strength = results['node_strength']
        node_degree = results['node_degree']
        clustering_coefficient = results['clustering_coefficient']
        local_efficiency_list = results['local_efficiency_list']
        
        for row_num, channel in enumerate(node_strength.keys(), 2):
            ws.cell(row=row_num, column=1, value=channel_labels[channel]).alignment = center_alignment
            ws.cell(row=row_num, column=2, value=node_strength[channel]).alignment = center_alignment
            ws.cell(row=row_num, column=3, value=node_degree[channel]).alignment = center_alignment
            ws.cell(row=row_num, column=4, value=clustering_coefficient[channel]).alignment = center_alignment
            ws.cell(row=row_num, column=5, value=local_efficiency_list[channel]).alignment = center_alignment
        
        # 写入单值指标并合并单元格
        characteristic_path_length = results['characteristic_path_length']
        global_efficiency = results['global_efficiency']
        local_efficiency = results['local_efficiency']
        
        # 合并 Characteristic Path Length
        ws.cell(row=2, column=6, value=characteristic_path_length).alignment = center_alignment
        ws.merge_cells(start_row=2, start_column=6, end_row=len(node_strength) + 1, end_column=6)
        
        # 合并 Global Efficiency
        ws.cell(row=2, column=7, value=global_efficiency).alignment = center_alignment
        ws.merge_cells(start_row=2, start_column=7, end_row=len(node_strength) + 1, end_column=7)
        
        # 合并 Local Efficiency
        ws.cell(row=2, column=8, value=local_efficiency).alignment = center_alignment
        ws.merge_cells(start_row=2, start_column=8, end_row=len(node_strength) + 1, end_column=8)
        # 提取节点强度值
        strength_values = list(node_strength.values())

        # 计算最小值和最大值
        min_strength = min(strength_values)
        max_strength = max(strength_values)
        if max_strength == min_strength:
            node_size_list = [60 for _ in strength_values]  # 默认值
        else:
            node_size_list = [((value - min_strength) * (100 - 20) / (max_strength - min_strength)) + 20 for value in strength_values]
        pdc_matrix=binary_matrix['weighted_pdc_matrix']
        # view=plotting.view_connectome(adjacency_matrix=band_pdc_matrix['weighted_pdc_matrix'] ,
        # node_coords=coords,
        # edge_cmap= 'Reds',          # 用来显示边的强度
        # edge_threshold=0, 
        # node_size=8,
        # linewidth=8,
        # title=(f'Brain Connectivity Map (PDC)in {band_name} Band '),
        # symmetric_cmap=False
        # )
        # view.open_in_browser()
    # 创建 figure
        plt.figure(figsize=(16, 6))  # 设置 figure 的大小

        # 绘制热图（第一个子图）
        plt.subplot(1, 2, 1)  # 1 行 2 列，第一个子图
        plt.imshow(pdc_matrix, cmap="OrRd", interpolation='nearest')
        plt.colorbar(label='PDC Value')
        plt.title(f'PDC Matrix in {band_name} Band')
        plt.xlabel('Target Channel')
        plt.ylabel('Source Channel')
        plt.xticks(np.arange(len(channel_labels)), channel_labels, rotation=45)
        plt.yticks(np.arange(len(channel_labels)), channel_labels)

        # 绘制脑网络连接图（第二个子图）
        plt.subplot(1, 2, 2)  # 1 行 2 列，第二个子图
        plotting.plot_connectome(
            adjacency_matrix=pdc_matrix,
            node_coords=coords,
            #node_color = '#00CCCC',
            node_size = node_size_list,# 标准化到 [20, 100]
            edge_threshold=0,  
            edge_vmin=pdc_matrix[pdc_matrix.nonzero()].min(),
            edge_vmax=pdc_matrix.max(),
            edge_cmap='rainbow',
            colorbar=True,
            edge_kwargs={'linewidth': 0.8},
            title=(f'Brain Connectivity Map (PDC) in {band_name} Band'),
            display_mode='ortho',  # 正交视图
            axes=plt.gca()  # 将图绘制到当前子图
        )

        # 调整布局
        plt.tight_layout()

        # 显示图像
        plt.show()


    # 删除默认创建的 sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    return wb

def eeg_pdc(root,raw):
    window = tk.Toplevel(root)
    window.withdraw()
    eeg_list, coords,channel_labels,fs=get_info(raw)
    # 保存为NumPy的`.npz`文件
    np.savez("my_list.npz", *eeg_list)
    wb=PDC_cal(eeg_list,coords,channel_labels,fs)
                # 弹出文件保存对话框
    excel_filename = asksaveasfilename(defaultextension=".xlsx",
                                        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                        title="Save Excel File")
                                        
    if excel_filename:  # 检查用户是否选择了文件名
        # 将 DataFrame 导出为 Excel 文件
        wb.save(excel_filename)

    else:
        messagebox.showwarning("Warning!","取消保存!")


